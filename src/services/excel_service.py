from openpyxl import load_workbook

from models.product import Product


class ExcelService:
    def read_products(self, file_path: str) -> list[Product]:
        workbook = load_workbook(file_path)
        sheet = workbook["MA-Produkte"]

        products = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            article_number, ean, mpn, name, manufacturer, category, ma_price, link = row

            product = Product(
                article_number=str(article_number),
                ean=str(ean or ""),
                mpn=str(mpn or ""),
                name=str(name),
                manufacturer=str(manufacturer or ""),
                category=str(category or ""),
                description="",
                ma_price=float(ma_price or 0),
                image_url="",
                link=str(link or ""),
            )

            products.append(product)

        return products 